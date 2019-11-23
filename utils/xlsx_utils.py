import os
import xlrd
from xlutils.copy import copy
import openpyxl

from utils.helpers import get_logger

log = get_logger(__name__)

class ExcelBase():
    def __init__(self, filepath, sheet_index=0, output_path=''):
        self.read_xlsx(filepath, sheet_index)
        self.output_path = output_path

    def output(self):
        self.wb.save(self.output_path)

class Xlsx(ExcelBase):

    def read_xlsx(self, filepath, sheet_index=0):
        self.wb = openpyxl.load_workbook(filepath)
        self.sheets = self.wb.get_sheet_names()
        self.sheet_r = self.wb.get_sheet_by_name(self.sheets[sheet_index])

    def contents(self, row_start=0, row_end=None, end_before_last_row=None):
        """

        :param row_start: 开始行
        :param row_end: 指定结束行
        :param end_before_last_row: 倒数第几行结束，最后剩下多少行要舍弃！
        :return:
        """
        #优先级：end_before_last_row > row_end，都不填默认是所有行。
        if not row_end:
            row_end = self.sheet_r.max_row
        if end_before_last_row:
            row_end = self.sheet_r.max_row - end_before_last_row

        for row in self.sheet_r.iter_rows(min_row=row_start, max_row=row_end):
            row_data = [col.value for col in row]
            log.debug(row_data)
            yield row_data

    def write_cell(self, row, col, content=''):
        self.sheet_r.cell(row, col, content)
        log.debug("write [{}] to row:{}, col:{}".format(content, row, col))
        return

class Xls(ExcelBase):

    def read_xlsx(self, filepath, sheet_index=0):
        self.rb = xlrd.open_workbook(filepath)
        self.wb = copy(self.rb)
        self.sheet_r = self.rb.sheet_by_index(sheet_index)
        self.sheet_w = self.wb.get_sheet(sheet_index)

    def contents(self, row_start=0, row_end=None, end_before_last_row=None):
        """

        :param row_start: 开始行
        :param row_end: 指定结束行
        :param end_before_last_row: 倒数第几行结束，最后剩下多少行要舍弃！
        :return:
        """
        #优先级：end_before_last_row > row_end，都不填默认是所有行。
        if not row_end:
            row_end = self.sheet_r.nrows
        if end_before_last_row:
            row_end = self.sheet_r.nrows - end_before_last_row

        for i in range(row_start, row_end, 1):
            yield self.sheet_r.row_values(i)

    def write_cell(self, row, col, content=''):
        self.sheet_w.write(row, col, content)
        log.debug("write [{}] to row:{}, col:{}".format(content, row, col))
        return


if __name__ == '__main__':
    xl = Xlsx("D:\workspace/Accounting/xlsx_model/记账凭证模板.xlsx", output_path="D:\workspace/Accounting/output/记账凭证模板.xlsx")
    # for i in xl.contents(row_start=0, end_before_last_row=1):
    #     print(i)
    xl.write_cell(6, 2, "hello")
    xl.output()